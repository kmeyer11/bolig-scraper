"""Regenerates a nicely formatted matches.xlsx snapshot from matches.csv.

matches.csv stays the append-only source of truth (simple, robust, no risk of
losing history if a run crashes mid-write). This just re-renders the full
history as a readable spreadsheet: bold header, sane column widths, currency/
size number formats, and real clickable hyperlinks — CSV alone can't do any
of that, which is why it looked messy in Excel.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HYPERLINK_FONT = Font(color="0563C1", underline="single")

# (kolonneoverskrift, csv-feltnavn, kolonnebredde)
COLUMNS = [
    ("Titel", "title", 42),
    ("Pris (kr/md)", "price_kr", 14),
    ("Størrelse", "size_m2", 12),
    ("Værelser", "rooms", 10),
    ("Ledig fra", "move_in", 16),
    ("Adresse", "address", 32),
    ("Site", "site", 12),
    ("Link", "url", 14),
    ("Fundet", "scraped_at", 17),
]


def _set_value(cell, field: str, value: str) -> None:
    if field == "price_kr":
        cell.value = float(value) if value else None
        cell.number_format = '#,##0 "kr"'
    elif field == "size_m2":
        cell.value = float(value) if value else None
        cell.number_format = '0.0 "m²"'
    elif field == "rooms":
        cell.value = float(value) if value else None
        cell.number_format = "0.#"
    elif field == "scraped_at":
        cell.value = value[:16].replace("T", " ") if value else ""
    elif field == "url":
        cell.value = "Se opslag" if value else ""
        if value:
            cell.hyperlink = value
            cell.font = HYPERLINK_FONT
    else:
        cell.value = value


def write_xlsx(csv_path: Union[str, Path], xlsx_path: Union[str, Path]) -> None:
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)
    if not csv_path.exists():
        return

    with csv_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    rows.sort(key=lambda r: r.get("scraped_at", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lejligheder"

    for col_idx, (header, _field, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_header, field, _width) in enumerate(COLUMNS, start=1):
            _set_value(ws.cell(row=row_idx, column=col_idx), field, row.get(field, ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 20

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
